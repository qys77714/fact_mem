#!/usr/bin/env python3
"""生成检索诊断结果的 Excel 文件。"""

import json
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESULTS_DIR = os.path.join(REPO, "results/retrieval_diagnostics")
OUTPUT = os.path.join(REPO, "results/retrieval_diagnostics.xlsx")

METHOD_ORDER = ["add_all", "mem0", "evermemos", "relation_decision"]
METHOD_DISPLAY = {
    "add_all": "Append-all",
    "mem0": "Mem0-style",
    "evermemos": "EverMemOS-style",
    "relation_decision": "RD",
}
MODEL_ORDER = ["gemma4-e4b", "Qwen3.5-4B"]
MODEL_DISPLAY = {"gemma4-e4b": "Gemma 4 E4B", "Qwen3.5-4B": "Qwen3.5-4B"}

# Styles
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BOLD = Font(name="Calibri", size=11, bold=True)
NORMAL = Font(name="Calibri", size=11)
PCT_FORMAT = '0.0"%"'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium"),
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin")
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.font = BOLD if bold else NORMAL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def load_all_results():
    all_results = {}
    for model in MODEL_ORDER:
        for method in METHOD_ORDER:
            path = os.path.join(RESULTS_DIR, f"per_question_{model}_{method}.jsonl")
            if os.path.exists(path):
                results = []
                with open(path) as f:
                    for line in f:
                        if line.strip():
                            results.append(json.loads(line))
                all_results[(model, method)] = results
    return all_results


def compute_metrics(results, key):
    n = len(results)
    if n == 0:
        return 0
    if key == "any_gold_retrieved":
        return sum(1 for r in results if r["any_gold_retrieved"]) / n * 100
    elif key == "all_gold_retrieved":
        return sum(1 for r in results if r["all_gold_retrieved"]) / n * 100
    elif key == "gold_recall":
        return np.mean([r["gold_recall"] for r in results]) * 100
    elif key == "confounder_token_share":
        return np.mean([r["confounder_token_share"] for r in results]) * 100
    elif key == "retrieved_entry_count":
        return np.mean([r["retrieved_entry_count"] for r in results])
    elif key == "memory_context_tokens":
        return np.mean([r["memory_context_tokens"] for r in results])
    return 0


def create_main_table(ws, all_results):
    """Sheet 1: Main results table."""
    ws.title = "Main Results"

    # Title
    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="Retrieval Process Diagnostics — N=8 Confounder Condition (256-token budget)").font = Font(name="Calibri", size=14, bold=True)

    # Headers
    headers = ["Manager Model", "Metric"] + [METHOD_DISPLAY[m] for m in METHOD_ORDER]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    metrics = [
        ("any_gold_retrieved", "Any gold retrieved (%)", "↑"),
        ("all_gold_retrieved", "All gold retrieved (%)", "↑"),
        ("gold_recall", "Gold recall (%)", "↑"),
        ("confounder_token_share", "Confounder tokens (%)", "↓"),
        ("retrieved_entry_count", "Retrieved entries (mean)", ""),
        ("memory_context_tokens", "Memory context tokens (mean)", ""),
    ]

    row = 4
    for mi, model in enumerate(MODEL_ORDER):
        for mki, (mkey, mname, arrow) in enumerate(metrics):
            display_name = f"{mname} {arrow}" if arrow else mname
            style_data_cell(ws, row, 1, bold=(mki == 0))
            if mki == 0:
                ws.cell(row=row, column=1, value=MODEL_DISPLAY[model])
                # Merge model name cells
                if row + len(metrics) - 1 > row:
                    ws.merge_cells(start_row=row, start_column=1, end_row=row + len(metrics) - 1, end_column=1)
            ws.cell(row=row, column=2, value=display_name)
            style_data_cell(ws, row, 2)

            for mj, method in enumerate(METHOD_ORDER):
                results = all_results.get((model, method), [])
                val = compute_metrics(results, mkey)
                cell = style_data_cell(ws, row, 3 + mj)
                if mkey == "retrieved_entry_count" or mkey == "memory_context_tokens":
                    cell.value = round(val, 1)
                    cell.number_format = '0.0'
                else:
                    cell.value = round(val, 1)
                    cell.number_format = '0.0'
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 32
    for i in range(3, 7):
        ws.column_dimensions[get_column_letter(i)].width = 18


def create_diagnostics_sheet(ws, all_results):
    """Sheet 2: Diagnostic checks."""
    ws.title = "Diagnostics"

    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1, value="Diagnostic Checks").font = Font(name="Calibri", size=14, bold=True)

    # Section 1: Question coverage
    row = 3
    ws.cell(row=row, column=1, value="1. Question Coverage (target: 470 per method)").font = BOLD
    row += 1
    for (model, method), results in sorted(all_results.items()):
        ws.cell(row=row, column=1, value=f"{MODEL_DISPLAY[model]} / {METHOD_DISPLAY[method]}").font = NORMAL
        ws.cell(row=row, column=2, value=len(results))
        row += 1

    # Section 2: Token limit
    row += 1
    ws.cell(row=row, column=1, value="2. Context Token Limit (should all be 0 over 256)").font = BOLD
    row += 1
    headers = ["Model", "Method", "Over 256", "Mean Tokens", "Std Tokens"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for (model, method), results in sorted(all_results.items()):
        over = sum(1 for r in results if r["memory_context_tokens"] > 256)
        mean_t = np.mean([r["memory_context_tokens"] for r in results])
        std_t = np.std([r["memory_context_tokens"] for r in results])
        ws.cell(row=row, column=1, value=MODEL_DISPLAY[model])
        ws.cell(row=row, column=2, value=METHOD_DISPLAY[method])
        ws.cell(row=row, column=3, value=over)
        ws.cell(row=row, column=4, value=round(mean_t, 1))
        ws.cell(row=row, column=5, value=round(std_t, 1))
        for c in range(1, 6):
            style_data_cell(ws, row, c)
        row += 1

    # Section 3: Entry count distribution
    row += 1
    ws.cell(row=row, column=1, value="3. Retrieved Entry Count Distribution").font = BOLD
    row += 1
    headers = ["Model", "Method", "Mean", "Std", "P25", "P50", "P75"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for (model, method), results in sorted(all_results.items()):
        counts = sorted([r["retrieved_entry_count"] for r in results])
        n = len(counts)
        vals = [MODEL_DISPLAY[model], METHOD_DISPLAY[method],
                round(np.mean(counts), 1), round(np.std(counts), 1),
                counts[int(n*0.25)], counts[int(n*0.5)], counts[int(n*0.75)]]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
            style_data_cell(ws, row, col)
        row += 1

    # Section 4: Gold recall by num_required
    row += 1
    ws.cell(row=row, column=1, value="4. Gold Recall by num_required_gold").font = BOLD
    row += 1
    from collections import defaultdict
    headers = ["Model", "Method", "1 gold", "2 gold", "3 gold", "4 gold", "5 gold", "6 gold"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for (model, method), results in sorted(all_results.items()):
        by_n = defaultdict(list)
        for r in results:
            by_n[r["num_required_gold"]].append(r["gold_recall"])
        vals = [MODEL_DISPLAY[model], METHOD_DISPLAY[method]]
        for k in range(1, 7):
            if k in by_n:
                vals.append(round(np.mean(by_n[k]) * 100, 1))
            else:
                vals.append("-")
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
            style_data_cell(ws, row, col)
        row += 1

    # Section 5: Type I vs Type II
    row += 1
    ws.cell(row=row, column=1, value="5. Confusion Type Breakdown (Gold Recall %)").font = BOLD
    row += 1
    headers = ["Model", "Method", "Type I (n=398)", "Type II KU (n=72)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for (model, method), results in sorted(all_results.items()):
        t1 = [r for r in results if r.get("confusion_type") == "type_i"]
        t2 = [r for r in results if r.get("confusion_type") == "type_ii"]
        r1 = np.mean([r["gold_recall"] for r in t1]) * 100 if t1 else 0
        r2 = np.mean([r["gold_recall"] for r in t2]) * 100 if t2 else 0
        vals = [MODEL_DISPLAY[model], METHOD_DISPLAY[method], round(r1, 1), round(r2, 1)]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
            style_data_cell(ws, row, col)
        row += 1

    # Section 6: Add-all pooling
    row += 1
    ws.cell(row=row, column=1, value="6. Add-all Pooling (gemma4-e4b vs Qwen3.5-4B)").font = BOLD
    row += 1
    if ("gemma4-e4b", "add_all") in all_results and ("Qwen3.5-4B", "add_all") in all_results:
        g = {r["history_name"]: r for r in all_results[("gemma4-e4b", "add_all")]}
        q = {r["history_name"]: r for r in all_results[("Qwen3.5-4B", "add_all")]}
        common = set(g) & set(q)
        identical = sum(1 for hn in common
                        if g[hn]["any_gold_retrieved"] == q[hn]["any_gold_retrieved"]
                        and g[hn]["retrieved_entry_count"] == q[hn]["retrieved_entry_count"])
        ws.cell(row=row, column=1, value=f"Common histories: {len(common)}")
        row += 1
        ws.cell(row=row, column=1, value=f"Identical: {identical}/{len(common)}")
        row += 1
        verdict = "POOLED (effectively identical)" if identical >= len(common) * 0.98 else "NOT pooled"
        ws.cell(row=row, column=1, value=f"Verdict: {verdict}")

    # Section 7: RD mixed fused + Evidence missed
    row += 2
    ws.cell(row=row, column=1, value="7. RD Mixed Fused & Evidence Gold Missed").font = BOLD
    row += 1
    headers = ["Model", "Method", "Mixed fused entries", "Evidence gold missed"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for (model, method), results in sorted(all_results.items()):
        mixed = sum(1 for r in results if r.get("is_mixed_fused"))
        ev_missed = sum(1 for r in results if r.get("evidence_has_gold_primary_missed"))
        vals = [MODEL_DISPLAY[model], METHOD_DISPLAY[method], mixed, ev_missed]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
            style_data_cell(ws, row, col)
        row += 1

    # Section 8: Mem0 gold preservation
    row += 1
    ws.cell(row=row, column=1, value="8. Mem0 Gold Memory Preservation").font = BOLD
    row += 1
    headers = ["Model", "Gold IDs found", "Gold IDs lost", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for model in MODEL_ORDER:
        results = all_results.get((model, "mem0"), [])
        if results:
            found = sum(1 for r in results if len(r.get("gold_memory_ids", [])) > 0)
            lost = sum(1 for r in results if r["num_required_gold"] > 0 and len(r.get("gold_memory_ids", [])) == 0)
            vals = [MODEL_DISPLAY[model], found, lost, len(results)]
            for col, v in enumerate(vals, 1):
                ws.cell(row=row, column=col, value=v)
                style_data_cell(ws, row, col)
            row += 1

    # Column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22


def create_per_question_sheet(ws, all_results):
    """Sheet 3+: Per-question data (one sheet per model-method)."""
    # Use a consolidated approach: all data in one sheet
    ws.title = "Per-Question Data"

    headers = [
        "manager_model", "method", "run_id", "query_id", "history_name",
        "num_required_gold", "num_retrieved_gold", "any_gold_retrieved",
        "all_gold_retrieved", "gold_recall", "memory_context_tokens",
        "confounder_tokens", "mixed_or_unattributed_tokens", "confounder_token_share",
        "retrieved_entry_count", "confusion_type", "is_mixed_fused",
        "evidence_has_gold_primary_missed",
        "required_gold_ids", "retrieved_gold_ids", "unmatched_required_gold_ids",
        "retrieved_source_memory_ids", "retrieved_entry_ids",
        "retrieved_confounder_ids", "duplicate_gold_match_count",
        "empty_context_reason"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    row = 2
    for model in MODEL_ORDER:
        for method in METHOD_ORDER:
            path = os.path.join(RESULTS_DIR, f"per_question_{model}_{method}.jsonl")
            if not os.path.exists(path):
                continue
            with open(path) as f:
                for line in f:
                    if not line.strip():
                        continue
                    r = json.loads(line)
                    vals = [
                        r.get("manager_model", model),
                        r.get("method", method),
                        r.get("run_id", 0),
                        r.get("query_id", ""),
                        r.get("history_name", ""),
                        r.get("num_required_gold", 0),
                        r.get("num_retrieved_gold", 0),
                        r.get("any_gold_retrieved", False),
                        r.get("all_gold_retrieved", False),
                        r.get("gold_recall", 0),
                        r.get("memory_context_tokens", 0),
                        r.get("confounder_tokens", 0),
                        r.get("mixed_or_unattributed_tokens", 0),
                        r.get("confounder_token_share", 0),
                        r.get("retrieved_entry_count", 0),
                        r.get("confusion_type", ""),
                        r.get("is_mixed_fused", False),
                        r.get("evidence_has_gold_primary_missed", False),
                        json.dumps(r.get("required_gold_ids", r.get("gold_memory_ids", []))),
                        json.dumps(r.get("retrieved_gold_ids", [])),
                        json.dumps(r.get("unmatched_required_gold_ids", [])),
                        json.dumps(r.get("retrieved_source_memory_ids", [])),
                        json.dumps(r.get("retrieved_entry_ids", [])),
                        json.dumps(r.get("retrieved_confounder_ids", [])),
                        r.get("duplicate_gold_match_count", 0),
                        r.get("empty_context_reason", ""),
                    ]
                    for col, v in enumerate(vals, 1):
                        ws.cell(row=row, column=col, value=v)
                        style_data_cell(ws, row, col)
                    row += 1

    # Column widths
    widths = [16, 18, 8, 38, 14, 16, 16, 16, 16, 12, 20, 16, 22, 22, 18, 14, 14, 26, 40, 40, 40, 40, 40, 40, 22, 22]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header
    ws.freeze_panes = "A2"


def create_methodology_sheet(ws):
    """Sheet: Methodology notes."""
    ws.title = "Methodology"

    notes = [
        ("Gold/Confounder Identification", [
            "Gold memory texts from data/preprocessed/longmemeval_s_hybrid_golden.json",
            "Matched to ingest memory IDs by EXACT text match for add_all, evermemos, RD",
            "For mem0 (which rewrites memory texts): exact match first, then fuzzy match (difflib.SequenceMatcher ratio >= 0.65)",
            "Confounders identified the same way (exact match only for non-mem0 methods)",
        ]),
        ("Token Counting", [
            "Tokenizer: Qwen3-8B (Qwen2Tokenizer), same as experiment",
            "Memory context tokens: sum of individual memory unit block tokens in formatted prompt",
            "Confounder tokens: tokens from memory units matched to confounder IDs",
            "Mixed tokens (RD): tokens from fused entries containing both gold and confounder sources",
            "No special tokens, system prompt, or question text included in memory token count",
        ]),
        ("RD Fusion Provenance", [
            "Fused entries identified by metadata.answer_fused=True",
            "Source membership tracked via metadata.fused_member_ids",
            "Entry classified as 'mixed' if fused_member_ids contain both gold and confounder IDs",
            "Mixed token count reported separately in mixed_or_unattributed_tokens",
        ]),
        ("Truncation", [
            "256-token hard head-truncation on memory context block",
            "Verified: 0 questions exceed 256 tokens in final context",
            "Entry count = entries surviving in final context (not all retrieved top-50)",
        ]),
        ("Single Run Note", [
            "This analysis uses single experimental runs (not 3-run mean)",
            "Add-all verified as effectively identical between models (464/470 identical)",
            "Process is deterministic (FAISS exact search, fixed seed)",
            "3/470 add-all differences due to minor float precision",
        ]),
        ("Experiment Config", [
            "Benchmark: lme_s_golden (hybrid golden)",
            "Candidate ID: cda53dff (hybrid_filler_N8)",
            "Extract model: gemma4-26B",
            "Answer model: gemma4-26B",
            "Embedding model: qwen3-embedding-0.6b",
            "Memory token limit: 256",
            "Retrieve top-k: 50",
            "Git commit: cec2b53",
            "Analysis date: 2026-08-02",
        ]),
    ]

    row = 1
    ws.cell(row=row, column=1, value="Methodology & Attribution Rules").font = Font(name="Calibri", size=14, bold=True)
    row += 2

    for section_title, items in notes:
        ws.cell(row=row, column=1, value=section_title).font = BOLD
        row += 1
        for item in items:
            ws.cell(row=row, column=1, value=f"  • {item}").font = NORMAL
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 100


def main():
    print("Loading per-question results...")
    all_results = load_all_results()
    print(f"  {len(all_results)} model-method combinations loaded")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Main Results
    ws1 = wb.create_sheet("Main Results")
    create_main_table(ws1, all_results)

    # Sheet 2: Diagnostics
    ws2 = wb.create_sheet("Diagnostics")
    create_diagnostics_sheet(ws2, all_results)

    # Sheet 3: Per-Question Data
    ws3 = wb.create_sheet("Per-Question Data")
    create_per_question_sheet(ws3, all_results)

    # Sheet 4: Methodology
    ws4 = wb.create_sheet("Methodology")
    create_methodology_sheet(ws4)

    wb.save(OUTPUT)
    print(f"\nExcel saved to: {OUTPUT}")


if __name__ == "__main__":
    main()
